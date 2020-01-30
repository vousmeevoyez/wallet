import string
from datetime import datetime, timedelta
from collections import OrderedDict

from app.api.models import Transaction, VirtualAccount, User, Payment, Wallet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def number_to_letter(number):
    d = dict(enumerate(string.ascii_uppercase, 0))
    return d[number]


def extract_transactions(start_time, end_time):
    transactions = Transaction.query.filter(
        Transaction.created_at.between(start_time, end_time)
    ).all()
    return transactions


class DailyTransactionReport:
    """ Daily transaction report excel """

    color = "f6e58d"
    color2 = "6ab04c"

    def __init__(self, data=None):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.data = data

    def insert_title(self):
        self.ws.merge_cells("A1:I1")
        self.ws.merge_cells("J1:R1")

        # put insert text
        self.ws["A1"].value = "Source"
        self.ws["A1"].alignment = Alignment(horizontal="center")
        self.ws["A1"].fill = PatternFill("solid", fgColor=self.color)
        self.ws["A1"].font = Font(size=16)

        self.ws["J1"].value = "Destination"
        self.ws["J1"].alignment = Alignment(horizontal="center")
        self.ws["J1"].fill = PatternFill("solid", fgColor=self.color2)
        self.ws["J1"].font = Font(size=16)

    @staticmethod
    def style_cell(styling, cell):
        """ styling cell """
        if "alignment" in styling:
            cell.alignment = Alignment(horizontal=styling["alignment"])

        if "color" in styling:
            cell.fill = PatternFill("solid", fgColor=styling["color"])

        if "font" in styling:
            cell.font = Font(size=styling["size"])

        if "currency" in styling:
            cell.number_format = "#,##0.00€"

    def insert_cell(self, row_position, values, style=None, column_position=1):
        """
            insert cell for particular row based on list of values
        """
        for value in values:
            cell = self.ws.cell(column=column_position, row=row_position, value=value)
            if style is not None:
                self.style_cell(style, cell)
            column_position += 1

    def adjust_column(self, values):
        """
            iterate through column and adjust width
        """
        for x in range(len(values)):
            self.ws.column_dimensions[number_to_letter(x)].width = 20

    def save(self):
        # create filename based on timestamp
        current_date = datetime.today().strftime("%Y%m%d_%H.%M")
        filename = "{} - Modanaku Daily Int Tx Report.xlsx".format(current_date)
        # set output directory
        output_folder = "data/reports/"
        filename = output_folder + filename
        self.wb.save(filename=filename)
        return filename

    def construct_header(self):
        # insert source and destination title
        self.insert_title()

        # define cells for title
        source_titles = [
            "Tx ID",
            "Tx Ref ID",
            "Source VA",
            "Phone Number",
            "Modanaku Type",
            "Tx Type",
            "Tx Date",
            "Tx Amount",
            "Balance after tx",
        ]
        destination_titles = source_titles

        # generate source column
        # insert column titles
        cell_style = {"alignment": "center", "color": self.color}
        self.insert_cell(row_position=2, values=source_titles, style=cell_style)

        # generate destination column
        # insert column titles
        cell_style = {"alignment": "center", "color": self.color2}
        self.insert_cell(
            row_position=2,
            values=destination_titles,
            style=cell_style,
            column_position=10,
        )
        # adjust column here
        self.adjust_column(values=source_titles + destination_titles)

    def _prettify_timestamp(self, timestamp):
        jkt_timestamp = timestamp + timedelta(hours=7)
        return jkt_timestamp.strftime("%Y-%m-%d %H:%M:%S")

    def _preprocess(self, row):
        # only output the ordered dict
        ordered_result = OrderedDict()

        # if its linked to other transaction then generate another
        if row.children != []:
            # maintain order
            child_trx = row.children[0]
            ordered_result["credit"] = {
                "trx_id": str(child_trx.id),
                "account_no": child_trx.wallet.virtual_account[0].account_no,
                "msisdn": child_trx.wallet.user.phone_ext
                + child_trx.wallet.user.phone_number,
                "wallet_type": child_trx.wallet.wallet_type,
                "trx_type": child_trx.transaction_type.key,
                "timestamp": self._prettify_timestamp(child_trx.created_at),
                "amount": child_trx.amount,
                "last_balance": child_trx.balance,
            }

        ordered_result["debit"] = {
            "trx_id": str(row.id),
            "account_no": row.wallet.virtual_accounts[0].account_no,
            "msisdn": row.wallet.user.phone_ext + row.wallet.user.phone_number,
            "wallet_type": row.wallet.label,
            "trx_type": row.transaction_type.key,
            "timestamp": self._prettify_timestamp(row.created_at),
            "amount": row.amount,
            "last_balance": row.balance,
        }

        return ordered_result

    def unpack(self, it):
        if isinstance(it, list):
            for sub_it in it:
                yield from self.unpack(sub_it)
        elif isinstance(it, dict):
            for value in it.values():
                yield from self.unpack(value)
        else:
            yield it

    def construct_body(self):
        # access body
        row_position = 3
        for row in self.data:
            # preprocess all row information here
            row = self._preprocess(row)
            # unpack into values
            values = self.unpack(row)
            self.insert_cell(row_position=row_position, values=values)
            row_position += 1

    def generate(self):
        self.construct_header()
        self.construct_body()
        filename = self.save()
        return filename
